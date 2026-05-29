#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
批量爬取多个城市地铁站数据
===========================
从 OpenStreetMap Overpass API 爬取多个城市的地铁站数据，
每个城市生成独立的 Excel 文件。

用法:
  python 爬取多个城市地铁站.py

依赖:
  pip install requests openpyxl
"""

import time
import json
import os
import re
import sys
from datetime import datetime

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ==================== 配置 ====================

OVERPASS_URL = 'https://overpass-api.de/api/interpreter'
REQUEST_INTERVAL = 3
COUNTRY_NAME = '中国'
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', '地铁xlsx')

HEADERS = {
    'Accept': '*/*',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Origin': 'https://overpass-turbo.eu',
    'Referer': 'https://overpass-turbo.eu/',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
}

# 查询模板
QUERY_URBAN_RAIL = "data=[out:json];area[name='{city}']->.a;(node['railway'='station']['station'='subway'](area.a);node['railway'='station']['station'='light_rail'](area.a);node['railway'='station']['station'='monorail'](area.a);way['railway'='station']['station'='subway'](area.a);way['railway'='station']['station'='light_rail'](area.a);way['railway'='station']['station'='monorail'](area.a););out body;"

QUERY_RELAXED = "data=[out:json];area[name='{city}']->.a;(node['railway'='station']['subway'='yes'](area.a);node['railway'='station']['station'='subway'](area.a);node['railway'='station']['station'='light_rail'](area.a);node['railway'='station']['station'='monorail'](area.a);way['railway'='station']['station'='subway'](area.a);way['railway'='station']['station'='light_rail'](area.a);way['railway'='station']['station'='monorail'](area.a););out body;"

# 城市列表
CITIES = [
    '北京市', '上海市', '广州市', '深圳市', '成都市', '杭州市', '武汉市', '南京市',
    '重庆市', '西安市', '苏州市', '天津市', '长沙市', '郑州市', '东莞市', '沈阳市',
    '青岛市', '合肥市', '佛山市', '宁波市', '昆明市', '大连市', '福州市', '厦门市',
    '哈尔滨市', '济南市', '温州市', '南宁市', '长春市', '泉州市', '石家庄市',
    '贵阳市', '南昌市', '金华市', '常州市', '珠海市', '嘉兴市', '南通市', '惠州市',
    '太原市', '中山市', '绍兴市', '乌鲁木齐市', '兰州市', '徐州市',
    '无锡市', '厦门市', '昆明市', '济南市',
]


# ==================== 核心函数 ====================

def query_city(city, relaxed=False):
    """查询 Overpass API"""
    template = QUERY_RELAXED if relaxed else QUERY_URBAN_RAIL
    payload = template.format(city=city)
    try:
        resp = requests.post(OVERPASS_URL, headers=HEADERS, data=payload, timeout=30)
        if resp.status_code != 200:
            return [], f'HTTP {resp.status_code}'
        return resp.json().get('elements', []), None
    except requests.Timeout:
        return [], '请求超时'
    except json.JSONDecodeError:
        return [], 'JSON解析失败'
    except Exception as e:
        return [], str(e)


def extract_stations(elements, city):
    """从 OSM 元素中提取车站信息"""
    results = []
    seen_keys = set()

    for elem in elements:
        tags = elem.get('tags') or {}
        if tags.get('usage') == 'freight':
            continue

        name = tags.get('name:zh') or tags.get('name') or ''
        if not name:
            continue
        if not name[-1] == "站":
            name += "站"

        name_en = tags.get('name:en', '')

        alias = tags.get('short_name', '')
        if not alias:
            alt = tags.get('alt_name', '')
            if alt:
                alias = re.split(r'[;；,，]', alt)[0].strip()

        lon = elem.get('lon')
        lat = elem.get('lat')
        if lon is None or lat is None:
            continue

        key = (city, name, round(float(lon), 6), round(float(lat), 6))
        if key in seen_keys:
            continue
        seen_keys.add(key)

        line = tags.get('line', '')
        is_transfer = 1 if line and re.search(r'[;；,，/]', line) else 0

        results.append({
            'country': COUNTRY_NAME,
            'city': city,
            'stationName': name,
            'stationNameEn': name_en,
            'stationAlias': alias,
            'longitude': round(float(lon), 6),
            'latitude': round(float(lat), 6),
            'isTransfer': is_transfer,
            'lineIds': '',
            'lineNames': line or tags.get('network', ''),
            'exitCount': 0,
            'hasToilet': 0,
            'stationType': 0,
            'openDate': '',
            'firstTime': '',
            'lastTime': '',
            'statusCode': 1,
            'extra': '',
        })

    return results


# ==================== Excel 生成 ====================

def save_to_excel(stations, output_path):
    """保存到 Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '地铁站数据'

    headers = ['国家', '城市', '站名', '英文名', '别称', '经度', '纬度',
               '换乘', '线路ID', '线路名', '出口数', '厕所', '类型',
               '开通日期', '首班', '末班', '状态码', '扩展']

    h_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    h_fill = PatternFill(start_color='409EFF', end_color='409EFF', fill_type='solid')
    h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font
        c.fill = h_fill
        c.alignment = h_align
        c.border = border

    d_font = Font(name='微软雅黑', size=10)
    d_align = Alignment(horizontal='center', vertical='center')

    for ri, s in enumerate(stations, 2):
        vals = [s['country'], s['city'], s['stationName'], s['stationNameEn'], s['stationAlias'],
                s['longitude'], s['latitude'], s['isTransfer'], s['lineIds'], s['lineNames'],
                s['exitCount'], s['hasToilet'], s['stationType'], s['openDate'], s['firstTime'],
                s['lastTime'], s['statusCode'], s['extra']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = d_font
            c.alignment = d_align
            c.border = border

    for i, w in enumerate([8, 10, 18, 20, 18, 12, 12, 6, 10, 20,
                           8, 6, 6, 12, 8, 8, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:R{len(stations) + 1}'
    wb.save(output_path)


def crawl_city(city):
    """爬取单个城市"""
    elems, err = query_city(city)
    if err:
        time.sleep(REQUEST_INTERVAL)
        elems, err = query_city(city, relaxed=True)
    if err:
        return [], err
    return extract_stations(elems, city), None


# ==================== 主流程 ====================

def main():
    print(f'开始批量爬取 {len(CITIES)} 个城市地铁站数据...', flush=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    total_stations = 0
    success_count = 0
    fail_count = 0

    for i, city in enumerate(CITIES, 1):
        print(f'\n[{i}/{len(CITIES)}] 正在爬取「{city}」...', flush=True)

        stations, err = crawl_city(city)

        if err:
            print(f'  ✗ 失败: {err}')
            fail_count += 1
            continue

        if not stations:
            print(f'  ○ 未找到地铁站数据')
            fail_count += 1
            continue

        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'地铁站数据_{city}_{timestamp}.xlsx'
        out_path = os.path.join(OUTPUT_DIR, filename)
        save_to_excel(stations, out_path)

        print(f'  ✓ {len(stations)} 个站点 → {filename}')
        total_stations += len(stations)
        success_count += 1

        time.sleep(REQUEST_INTERVAL)

    print(f'\n{"="*50}')
    print(f'爬取完成!')
    print(f'  成功: {success_count} 个城市')
    print(f'  失败: {fail_count} 个城市')
    print(f'  总站点数: {total_stations}')
    print(f'  输出目录: {OUTPUT_DIR}')


if __name__ == '__main__':
    main()

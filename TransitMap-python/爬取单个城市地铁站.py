#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
爬取单个城市地铁站数据
======================
从 OpenStreetMap Overpass API 爬取指定城市的地铁站数据，
每个城市生成独立的 Excel 文件。

用法:
  python 爬取单个城市地铁站.py 北京市
  python 爬取单个城市地铁站.py 东莞市
  python 爬取单个城市地铁站.py 香港

Excel 列:
  国家|城市|站名|英文名|别称|经度|纬度|换乘|线路ID|线路名|出口数|厕所|类型|开通日期|首班|末班|状态码|扩展

依赖:
  pip install requests openpyxl
"""

import time
import json
import os
import re
import sys
from datetime import datetime

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ==================== 配置 ====================

OVERPASS_URL = 'https://overpass-api.de/api/interpreter'
REQUEST_INTERVAL = 2
COUNTRY_NAME = '中国'
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', '地铁xlsx')

HEADERS = {
    'Accept': '*/*',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Origin': 'https://overpass-turbo.eu',
    'Referer': 'https://overpass-turbo.eu/',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
}

# 查询模板：地铁/轻轨/单轨 站
QUERY_URBAN_RAIL = "data=[out:json];area[name='{city}']->.a;(node['railway'='station']['station'='subway'](area.a);node['railway'='station']['station'='light_rail'](area.a);node['railway'='station']['station'='monorail'](area.a);way['railway'='station']['station'='subway'](area.a);way['railway'='station']['station'='light_rail'](area.a);way['railway'='station']['station'='monorail'](area.a););out body;"

# 宽松模板：包含 subway=yes 的常规车站
QUERY_RELAXED = "data=[out:json];area[name='{city}']->.a;(node['railway'='station']['subway'='yes'](area.a);node['railway'='station']['station'='subway'](area.a);node['railway'='station']['station'='light_rail'](area.a);node['railway'='station']['station'='monorail'](area.a);way['railway'='station']['station'='subway'](area.a);way['railway'='station']['station'='light_rail'](area.a);way['railway'='station']['station'='monorail'](area.a););out body;"


# ==================== 核心函数 ====================

def query_city(city, relaxed=False):
    """查询 Overpass API，返回 (元素列表, 错误信息)"""
    template = QUERY_RELAXED if relaxed else QUERY_URBAN_RAIL
    payload = template.format(city=city)
    try:
        resp = requests.post(OVERPASS_URL, headers=HEADERS, data=payload, timeout=30)
        if resp.status_code != 200:
            return [], f'HTTP {resp.status_code}'
        return resp.json().get('elements', []), None
    except requests.Timeout:
        return [], '请求超时'
    except json.JSONDecodeError:
        return [], 'JSON解析失败'
    except Exception as e:
        return [], str(e)


def extract_stations(elements, city):
    """从 OSM 元素列表中提取车站信息"""
    results = []
    seen_keys = set()

    for elem in elements:
        tags = elem.get('tags') or {}
        # 跳过货运站
        if tags.get('usage') == 'freight':
            continue

        # 站名
        name = tags.get('name:zh') or tags.get('name') or ''
        if not name:
            continue
        if not name[-1] =="站":
            name+="站"
        # 英文名
        name_en = tags.get('name:en', '')

        # 别称
        alias = tags.get('short_name', '')
        if not alias:
            alt = tags.get('alt_name', '')
            if alt:
                alias = re.split(r'[;；,，]', alt)[0].strip()

        lon = elem.get('lon')
        lat = elem.get('lat')
        if lon is None or lat is None:
            continue

        # 去重键 (城市 + 站名 + 经纬度前6位)
        key = (city, name, round(float(lon), 6), round(float(lat), 6))
        if key in seen_keys:
            continue
        seen_keys.add(key)

        # 换乘判断
        line = tags.get('line', '')
        is_transfer = 1 if line and re.search(r'[;；,，/]', line) else 0

        results.append({
            'country': COUNTRY_NAME,
            'city': city,
            'stationName': name,
            'stationNameEn': name_en,
            'stationAlias': alias,
            'longitude': round(float(lon), 6),
            'latitude': round(float(lat), 6),
            'isTransfer': is_transfer,
            'lineIds': '',
            'lineNames': line or tags.get('network', ''),
            'exitCount': 0,
            'hasToilet': 0,
            'stationType': 0,
            'openDate': '',
            'firstTime': '',
            'lastTime': '',
            'statusCode': 1,
            'extra': '',
        })

    return results


# ==================== Excel 生成 ====================

def save_to_excel(stations, output_path):
    """保存车站列表到 Excel 文件"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '地铁站数据'

    headers = ['国家', '城市', '站名', '英文名', '别称', '经度', '纬度',
               '换乘', '线路ID', '线路名', '出口数', '厕所', '类型',
               '开通日期', '首班', '末班', '状态码', '扩展']

    h_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    h_fill = PatternFill(start_color='409EFF', end_color='409EFF', fill_type='solid')
    h_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font
        c.fill = h_fill
        c.alignment = h_align
        c.border = border

    d_font = Font(name='微软雅黑', size=10)
    d_align = Alignment(horizontal='center', vertical='center')

    for ri, s in enumerate(stations, 2):
        vals = [s['country'], s['city'], s['stationName'], s['stationNameEn'], s['stationAlias'],
                s['longitude'], s['latitude'], s['isTransfer'], s['lineIds'], s['lineNames'],
                s['exitCount'], s['hasToilet'], s['stationType'], s['openDate'], s['firstTime'],
                s['lastTime'], s['statusCode'], s['extra']]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = d_font
            c.alignment = d_align
            c.border = border

    for i, w in enumerate([8, 10, 18, 20, 18, 12, 12, 6, 10, 20,
                           8, 6, 6, 12, 8, 8, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:R{len(stations) + 1}'
    wb.save(output_path)


def crawl_city(city):
    """爬取单个城市，返回 (车站列表, 错误信息)"""
    elems, err = query_city(city)
    if err:
        # 失败后尝试宽松查询
        time.sleep(REQUEST_INTERVAL)
        elems, err = query_city(city, relaxed=True)
    if err:
        return [], err

    stations = extract_stations(elems, city)
    return stations, None


# ==================== 主流程 ====================

def main():
    # if len(sys.argv) < 2:
    #     print('用法: python 爬取单个城市地铁站.py <城市名>')
    #     print('示例: python 爬取单个城市地铁站.py 北京市')
    #     sys.exit(1)

    city = '南昌市'

    print(f'正在爬取「{city}」地铁站数据...', flush=True)
    os.makedirs(OUTPUT_DIR, exist_ok=True)

    stations, err = crawl_city(city)

    if err:
        print(f'✗ 爬取失败: {err}')
        sys.exit(1)

    if not stations:
        print(f'○ 「{city}」未找到地铁站数据（可能 OSM 中暂无地铁站标记）')
        sys.exit(0)

    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    filename = f'地铁站数据_{city}_{timestamp}.xlsx'
    out_path = os.path.join(OUTPUT_DIR, filename)
    save_to_excel(stations, out_path)

    print(f'✓ 爬取完成: {len(stations)} 个站点')
    print(f'  Excel: {out_path}')
    print(f'  可直接在管理后台「地铁站管理→批量导入」上传')


if __name__ == '__main__':
    main()

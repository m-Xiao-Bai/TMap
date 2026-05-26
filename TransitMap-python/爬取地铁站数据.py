#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
地铁站数据爬取脚本
===================
从 OpenStreetMap Overpass API 爬取中国主要城市地铁站数据，
生成符合 TransitMap 管理后台「批量导入」格式的 Excel 文件。

用法:
  python 爬取地铁站数据.py                    # 爬取全部城市
  python 爬取地铁站数据.py 北京市              # 爬取单个城市
  python 爬取地铁站数据.py 北京市 上海市 广州市 # 爬取多个城市
  python 爬取地铁站数据.py --list              # 查看可用城市列表

Excel 列说明（与 MetroStationManageController.parseExcel 对应）:
  0: 国家(名称)   1: 城市(名称)   2: 站名      3: 英文名     4: 别称
  5: 经度        6: 纬度         7: 换乘标志   8: 线路ID     9: 线路名
  10: 出口数     11: 厕所        12: 类型      13: 开通日期   14: 首班
  15: 末班       16: 状态码      17: 扩展数据

注意事项:
  - 国家/城市列使用名称而非数据库 ID，后端导入时会自动解析
  - 爬取前请确保目标国家/城市已在后台「国家管理」「城市管理」中添加

依赖安装:
  pip install requests openpyxl
"""

import time
import json
import os
import re
import sys
from datetime import datetime

import requests
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


# ==================== 配置区域 ====================

# Overpass API 端点（可切换镜像）
OVERPASS_URL = 'https://overpass-api.de/api/interpreter'
# OVERPASS_URL = 'https://overpass.kumi.systems/api/interpreter'

# 请求间隔（秒），避免被限流
REQUEST_INTERVAL = 3

# 国家名称（所有城市均属于中国）
COUNTRY_NAME = '中国'

# 输出目录
OUTPUT_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'data', '地铁xlsx')

# ==================== 城市列表 ====================

CITIES = [
    # 华北
    '北京市', '天津市', '石家庄市', '太原市', '呼和浩特市',
    # 东北
    '沈阳市', '大连市', '长春市', '哈尔滨市',
    # 华东
    '上海市', '南京市', '无锡市', '徐州市', '常州市', '苏州市',
    '杭州市', '宁波市', '温州市', '嘉兴市', '绍兴市',
    '合肥市', '福州市', '厦门市', '南昌市', '济南市', '青岛市',
    # 中南
    '郑州市', '武汉市', '长沙市', '广州市', '深圳市', '珠海市',
    '佛山市', '东莞市', '南宁市', '海口市',
    # 西南
    '重庆市', '成都市', '贵阳市', '昆明市', '拉萨市',
    # 西北
    '西安市', '兰州市', '西宁市', '银川市', '乌鲁木齐市',
    # 港澳台
    '香港', '澳门', '台北',
]

# ==================== Overpass 查询 ====================

# 查询模板 - 地铁站 (subway)，与测试脚本一致的格式
QUERY_SUBWAY = "data=[out:json];area[name='{city}']->.a;(  node['railway'='station']['station'='subway'](area.a);  way['railway'='station']['station'='subway'](area.a););out body;"

# 宽松模板 - 包含轻轨/单轨等
QUERY_RELAXED = "data=[out:json];area[name='{city}']->.a;(  node['railway'='station']['station'='subway'](area.a);  node['railway'='station']['station'='light_rail'](area.a);  node['railway'='station']['station'='monorail'](area.a);  way['railway'='station']['station'='subway'](area.a););out body;"

HEADERS = {
    'Accept': '*/*',
    'Accept-Language': 'zh-CN,zh;q=0.9',
    'Connection': 'keep-alive',
    'Content-Type': 'application/x-www-form-urlencoded; charset=UTF-8',
    'Origin': 'https://overpass-turbo.eu',
    'Referer': 'https://overpass-turbo.eu/',
    'Sec-Fetch-Dest': 'empty',
    'Sec-Fetch-Mode': 'cors',
    'Sec-Fetch-Site': 'cross-site',
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/143.0.0.0 Safari/537.36',
    'sec-ch-ua': '"Google Chrome";v="143", "Chromium";v="143", "Not A(Brand";v="24"',
    'sec-ch-ua-mobile': '?0',
    'sec-ch-ua-platform': '"Windows"',
}


def query_overpass(city, use_relaxed=False):
    """查询 Overpass API 获取城市地铁站数据，返回 (元素列表, 错误信息)"""
    template = QUERY_RELAXED if use_relaxed else QUERY_SUBWAY
    # 注意：data 必须作为原始字符串发送（带 data= 前缀），与测试脚本一致
    data_str = template.format(city=city)

    try:
        resp = requests.post(OVERPASS_URL, headers=HEADERS, data=data_str, timeout=30)
        if resp.status_code != 200:
            return [], f'HTTP {resp.status_code}'
        data = resp.json()
        return data.get('elements', []), None
    except requests.Timeout:
        return [], '请求超时'
    except json.JSONDecodeError:
        return [], 'JSON解析失败'
    except Exception as e:
        return [], str(e)


def extract_station(elem, city):
    """从 OSM element 中提取车站信息字典"""
    tags = elem.get('tags') or {}
    elem_type = elem.get('type', '')

    # 跳过货运站
    if tags.get('usage') == 'freight':
        return None

    # 站名（优先中文）
    name = tags.get('name:zh') or tags.get('name') or ''
    if not name:
        return None

    # 英文名
    name_en = tags.get('name:en', '')

    # 别称
    alias = tags.get('short_name', '')
    if not alias:
        alt = tags.get('alt_name', '')
        if alt:
            parts = re.split(r'[;；,，]', alt)
            alias = parts[0].strip()

    # 经纬度
    lon = elem.get('lon')
    lat = elem.get('lat')
    if lon is None or lat is None:
        return None

    # 换乘判断：line 标签包含多线路，或站名含换乘标记
    line = tags.get('line', '')
    is_transfer = 1 if (line and re.search(r'[;；,，/]', line)) else 0

    # 线路名
    line_names = line or tags.get('network', '')

    return {
        'country': COUNTRY_NAME,
        'city': city,
        'stationName': name,
        'stationNameEn': name_en,
        'stationAlias': alias,
        'longitude': round(float(lon), 6),
        'latitude': round(float(lat), 6),
        'isTransfer': is_transfer,
        'lineIds': '',
        'lineNames': line_names,
        'exitCount': 0,
        'hasToilet': 0,
        'stationType': 0,
        'openDate': '',
        'firstTime': '',
        'lastTime': '',
        'statusCode': 1,
        'extra': '',
    }


def deduplicate(stations):
    """基于 城市+站名+经纬度 去重"""
    seen = set()
    result = []
    for s in stations:
        key = (s['city'], s['stationName'],
               s['longitude'], s['latitude'],
               s.get('osmid', ''))
        # 同时也对同名同城去重（不同经纬度但站名相同 → 合并）
        key2 = (s['city'], s['stationName'])
        if key not in seen and key2 not in seen:
            seen.add(key)
            seen.add(key2)
            result.append(s)
    return result


# ==================== Excel 生成 ====================

def create_excel(stations, output_path):
    """
    生成符合导入格式的 Excel（18列）
    列: 国家|城市|站名|英文名|别称|经度|纬度|换乘|线路ID|线路名|出口数|厕所|类型|开通日期|首班|末班|状态码|扩展
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '地铁站数据'

    headers = [
        '国家', '城市', '站名', '英文名', '别称',
        '经度', '纬度', '换乘', '线路ID', '线路名',
        '出口数', '厕所', '类型', '开通日期', '首班',
        '末班', '状态码', '扩展',
    ]

    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='409EFF', end_color='409EFF', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'),
    )

    # 写入表头
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = header_font
        c.fill = header_fill
        c.alignment = header_align
        c.border = thin_border

    data_font = Font(name='微软雅黑', size=10)
    data_align = Alignment(horizontal='center', vertical='center')

    for row_idx, s in enumerate(stations, 2):
        vals = [
            s['country'], s['city'], s['stationName'], s['stationNameEn'], s['stationAlias'],
            s['longitude'], s['latitude'], s['isTransfer'], s['lineIds'], s['lineNames'],
            s['exitCount'], s['hasToilet'], s['stationType'], s['openDate'], s['firstTime'],
            s['lastTime'], s['statusCode'], s['extra'],
        ]
        for col, v in enumerate(vals, 1):
            c = ws.cell(row=row_idx, column=col, value=v)
            c.font = data_font
            c.alignment = data_align
            c.border = thin_border

    # 列宽
    for i, w in enumerate([8, 10, 18, 20, 18, 12, 12, 6, 10, 20,
                           8, 6, 6, 12, 8, 8, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f'A1:R{len(stations) + 1}'

    wb.save(output_path)


# ==================== 主流程 ====================

def parse_cities():
    """解析命令行参数，返回待爬取的城市列表"""
    args = sys.argv[1:]

    if not args:
        # 无参数 → 爬取全部
        return CITIES

    # 过滤出在 CITIES 中的有效城市名，同时保留原始输入顺序
    requested = []
    unknown = []
    for name in args:
        name = name.strip()
        if name in CITIES:
            requested.append(name)
        else:
            unknown.append(name)

    if unknown:
        print(f'⚠ 以下城市不在预设列表中，将被忽略: {unknown}')
        print(f'   可用城市可通过 --list 查看')

    if not requested:
        print('❌ 未匹配到有效城市')
        print_available_cities()
        sys.exit(1)

    return requested


def print_available_cities():
    """打印可用城市列表"""
    print(f'\n可用城市 ({len(CITIES)} 个):')
    for region, cities_in_region in [
        ('华北', ['北京市', '天津市', '石家庄市', '太原市', '呼和浩特市']),
        ('东北', ['沈阳市', '大连市', '长春市', '哈尔滨市']),
        ('华东', ['上海市', '南京市', '无锡市', '徐州市', '常州市', '苏州市',
                  '杭州市', '宁波市', '温州市', '嘉兴市', '绍兴市',
                  '合肥市', '福州市', '厦门市', '南昌市', '济南市', '青岛市']),
        ('中南', ['郑州市', '武汉市', '长沙市', '广州市', '深圳市', '珠海市',
                  '佛山市', '东莞市', '南宁市', '海口市']),
        ('西南', ['重庆市', '成都市', '贵阳市', '昆明市', '拉萨市']),
        ('西北', ['西安市', '兰州市', '西宁市', '银川市', '乌鲁木齐市']),
        ('港澳台', ['香港', '澳门', '台北']),
    ]:
        print(f'  [{region}] ' + ', '.join(cities_in_region))


def main():
    print('=' * 60)
    print('  地铁站数据爬取工具')
    print(f'  开始时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}')

    # 解析命令行参数
    if '--list' in sys.argv or '-l' in sys.argv:
        print_available_cities()
        return

    cities = parse_cities()
    print(f'  目标城市: {len(cities)} 个 → {", ".join(cities)}')
    print('=' * 60)

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    all_stations = []
    failed = []
    empty = []

    for idx, city in enumerate(cities, 1):
        print(f'\n[{idx:2d}/{len(cities)}] {city} ...', end=' ', flush=True)

        elems, err = query_overpass(city)
        if err:
            print(f'失败({err}), 尝试宽松查询...', end=' ', flush=True)
            time.sleep(REQUEST_INTERVAL)
            elems, err = query_overpass(city, use_relaxed=True)

        if err:
            print(f'✗ {err}')
            failed.append((city, err))
            continue

        city_stations = []
        for e in elems:
            s = extract_station(e, city)
            if s:
                city_stations.append(s)

        if city_stations:
            print(f'✓ {len(city_stations)} 站')
            all_stations.extend(city_stations)
        else:
            print('○ 无地铁站数据')
            empty.append(city)

        if idx < len(cities):
            time.sleep(REQUEST_INTERVAL)

    print(f'\n{"=" * 60}')
    print(f'原始采集: {len(all_stations)} 条')

    all_stations = deduplicate(all_stations)
    print(f'去重后:   {len(all_stations)} 条')

    if all_stations:
        ts = datetime.now().strftime('%Y%m%d_%H%M%S')
        city_suffix = f'_{len(cities)}城' if len(cities) <= 5 else f'_{len(cities)}城'
        filename = f'地铁站数据_{ts}{city_suffix}.xlsx'
        out_path = os.path.join(OUTPUT_DIR, filename)
        create_excel(all_stations, out_path)
        print(f'\n✅ Excel 已生成: {out_path}')
        print(f'   共 {len(all_stations)} 条记录，可直接在管理后台导入')

    print(f'\n📊 统计: 成功={len(cities)-len(failed)-len(empty)}, '
          f'失败={len(failed)}, 无数据={len(empty)}')
    if failed:
        print('失败城市:')
        for c, e in failed:
            print(f'  - {c}: {e}')
    if empty:
        print('无数据城市:')
        for c in empty:
            print(f'  - {c}')
    print('=' * 60)


if __name__ == '__main__':
    main()

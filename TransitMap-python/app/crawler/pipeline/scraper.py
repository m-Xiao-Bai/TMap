"""
爬虫管线 — Step 1: 数据爬取

从 OpenStreetMap Overpass API 爬取指定城市的地铁站数据。
城市名由前端传入。
爬取完成后保存 Excel 文件。
"""

import asyncio
import json
import logging
import os
import re
import time
from datetime import datetime

import httpx

from app.crawler.sources.base import CrawlResult, ScrapedLine, ScrapedStation
from app.crawler.progress_tracker import progress_tracker

logger = logging.getLogger("tmap-python.crawler.pipeline.scraper")

# Overpass API 配置
OVERPASS_URL = "https://overpass-api.de/api/interpreter"
REQUEST_INTERVAL = 2

HEADERS = {
    "Accept": "*/*",
    "Accept-Language": "zh-CN,zh;q=0.9",
    "Connection": "keep-alive",
    "Content-Type": "application/x-www-form-urlencoded; charset=UTF-8",
    "Origin": "https://overpass-turbo.eu",
    "Referer": "https://overpass-turbo.eu/",
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36",
}

# 查询模板：地铁/轻轨/单轨 站（与原脚本一致）
QUERY_URBAN_RAIL = "data=[out:json];area[name='{city}']->.a;(node['railway'='station']['station'='subway'](area.a);node['railway'='station']['station'='light_rail'](area.a);node['railway'='station']['station'='monorail'](area.a);way['railway'='station']['station'='subway'](area.a);way['railway'='station']['station'='light_rail'](area.a);way['railway'='station']['station'='monorail'](area.a););out body;"

# 宽松模板：包含 subway=yes 的常规车站（与原脚本一致）
QUERY_RELAXED = "data=[out:json];area[name='{city}']->.a;(node['railway'='station']['subway'='yes'](area.a);node['railway'='station']['station'='subway'](area.a);node['railway'='station']['station'='light_rail'](area.a);node['railway'='station']['station'='monorail'](area.a);way['railway'='station']['station'='subway'](area.a);way['railway'='station']['station'='light_rail'](area.a);way['railway'='station']['station'='monorail'](area.a););out body;"

# Excel 输出目录
EXCEL_OUTPUT_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))),
    "data", "地铁xlsx"
)


async def scrape(task_id: str, city_name: str, sources: list[str] = None) -> list[CrawlResult]:
    """
    从 OSM Overpass API 爬取城市地铁站数据。

    Args:
        task_id: 任务 ID（用于进度追踪）
        city_name: 城市名称（前端传入，如 "南宁市"）
        sources: 忽略此参数，固定使用 OSM

    Returns:
        包含单个 CrawlResult 的列表
    """
    await progress_tracker.update(task_id, 5, "scraping", f"开始爬取「{city_name}」...")

    result = CrawlResult(source="osm", city_name=city_name)

    try:
        # 严格查询
        elements, err = await _query_city(city_name, relaxed=False)

        if err:
            # 失败后尝试宽松查询
            logger.info(f"严格查询失败({err})，尝试宽松查询...")
            await asyncio.sleep(REQUEST_INTERVAL)
            elements, err = await _query_city(city_name, relaxed=True)

        if err:
            result.success = False
            result.error = err
            logger.error(f"OSM 爬取失败: {err}")
            return [result]

        # 提取站点和线路
        stations, lines = _extract_data(elements, city_name)
        result.stations = stations
        result.lines = lines

        logger.info(f"OSM 爬取完成: {len(lines)} 条线路, {len(stations)} 个站点")

        # 保存 Excel
        try:
            _save_to_excel(city_name, stations)
        except Exception:
            pass

    except Exception as e:
        result.success = False
        result.error = str(e)

    await progress_tracker.update(
        task_id, 15, "scraping",
        f"爬取完成: {len(result.stations)} 个站点",
    )

    return [result]


async def _query_city(city: str, relaxed: bool = False) -> tuple[list[dict], str | None]:
    """
    查询 Overpass API。

    与原脚本 query_city() 逻辑一致。

    Returns:
        (元素列表, 错误信息)
    """
    template = QUERY_RELAXED if relaxed else QUERY_URBAN_RAIL
    payload = template.format(city=city)

    try:
        async with httpx.AsyncClient(timeout=30) as client:
            resp = await client.post(
                OVERPASS_URL,
                content=payload,
                headers=HEADERS,
            )

            if resp.status_code != 200:
                return [], f"HTTP {resp.status_code}"

            data = resp.json()
            return data.get("elements", []), None

    except httpx.TimeoutException:
        return [], "请求超时"
    except json.JSONDecodeError:
        return [], "JSON解析失败"
    except Exception as e:
        return [], str(e)


def _extract_data(elements: list[dict], city: str) -> tuple[list[ScrapedStation], list[ScrapedLine]]:
    """
    从 OSM 元素中提取站点和线路数据。

    与原脚本 extract_stations() 逻辑一致。
    """
    stations = []
    lines_map = {}
    seen_keys = set()

    for elem in elements:
        tags = elem.get("tags") or {}

        # 跳过货运站
        if tags.get("usage") == "freight":
            continue

        # 站名
        name = tags.get("name:zh") or tags.get("name") or ""
        if not name:
            continue
        if not name.endswith("站"):
            name += "站"

        # 英文名
        name_en = tags.get("name:en", "")

        # 别称
        alias = tags.get("short_name", "")
        if not alias:
            alt = tags.get("alt_name", "")
            if alt:
                alias = re.split(r"[;；,，]", alt)[0].strip()

        # 坐标
        lon = elem.get("lon")
        lat = elem.get("lat")
        if lon is None or lat is None:
            # way 类型需要计算中心点
            if elem.get("type") == "way" and "center" in elem:
                lon = elem["center"].get("lon")
                lat = elem["center"].get("lat")
            if lon is None or lat is None:
                continue

        # 去重键 (城市 + 站名 + 经纬度前6位)
        key = (city, name, round(float(lon), 6), round(float(lat), 6))
        if key in seen_keys:
            continue
        seen_keys.add(key)

        # 线路信息（添加城市前缀，如 "1号线" → "南宁1号线"）
        line = tags.get("line", "")
        is_transfer = bool(line and re.search(r"[;；,，/]", line))

        # 拆分多线路
        line_names = []
        if line:
            raw_lines = re.split(r"[;；,，/]", line)
            raw_lines = [ln.strip() for ln in raw_lines if ln.strip()]
            # 给每条线路名添加城市前缀
            city_short = city.replace("市", "").replace("地区", "")
            for ln in raw_lines:
                # 如果线路名已经包含城市前缀则不重复添加
                if ln.startswith(city_short):
                    line_names.append(ln)
                else:
                    line_names.append(f"{city_short}{ln}")

        # 注册线路
        for ln in line_names:
            if ln not in lines_map:
                lines_map[ln] = ScrapedLine(name=ln)
            if name not in lines_map[ln].stations:
                lines_map[ln].stations.append(name)

        # osmid
        osmid = elem.get("id", 0)

        # 站点的线路名（添加城市前缀）
        station_line_name = ";".join(line_names) if line_names else ""

        stations.append(ScrapedStation(
            name=name,
            name_en=name_en,
            alias=alias,
            line_name=station_line_name,
            is_transfer=is_transfer,
            lat=round(float(lat), 6),
            lng=round(float(lon), 6),
            osmid=osmid,
        ))

    return stations, list(lines_map.values())


def _save_to_excel(city_name: str, stations: list[ScrapedStation]) -> str:
    """
    保存站点数据到 Excel 文件。

    格式与原脚本 save_to_excel() 完全一致。
    """
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter

    os.makedirs(EXCEL_OUTPUT_DIR, exist_ok=True)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "地铁站数据"

    # 表头（与原脚本一致）
    headers = ["国家", "城市", "站名", "英文名", "别称", "经度", "纬度",
               "换乘", "线路ID", "线路名", "出口数", "厕所", "类型",
               "开通日期", "首班", "末班", "状态码", "扩展"]

    h_font = Font(name="微软雅黑", bold=True, size=11, color="FFFFFF")
    h_fill = PatternFill(start_color="409EFF", end_color="409EFF", fill_type="solid")
    h_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = h_font
        c.fill = h_fill
        c.alignment = h_align
        c.border = border

    d_font = Font(name="微软雅黑", size=10)
    d_align = Alignment(horizontal="center", vertical="center")

    for ri, s in enumerate(stations, 2):
        vals = [
            "中国",                          # 国家
            city_name,                       # 城市
            s.name,                          # 站名
            s.name_en,                       # 英文名
            s.alias,                         # 别称
            s.lng,                           # 经度
            s.lat,                           # 纬度
            1 if s.is_transfer else 0,       # 换乘
            "",                              # 线路ID
            s.line_name,                     # 线路名
            0,                               # 出口数
            0,                               # 厕所
            0,                               # 类型
            "",                              # 开通日期
            "",                              # 首班
            "",                              # 末班
            1,                               # 状态码
            "",                              # 扩展
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.font = d_font
            c.alignment = d_align
            c.border = border

    # 列宽（与原脚本一致）
    for i, w in enumerate([8, 10, 18, 20, 18, 12, 12, 6, 10, 20,
                           8, 6, 6, 12, 8, 8, 8, 10], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:R{len(stations) + 1}"

    # 保存
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"地铁站数据_{city_name}_{timestamp}.xlsx"
    filepath = os.path.join(EXCEL_OUTPUT_DIR, filename)
    wb.save(filepath)

    return filepath
